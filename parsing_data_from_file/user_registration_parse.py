# config_reader.py

from openpyxl import load_workbook


class ParseDataExcel:
    def __init__(self, filename):
        self.filename = filename

    def get_tests_to_run(self, sheet_name=None, start_row=2, end_row=None):
        workbook = load_workbook(filename=self.filename)
        if sheet_name:
            sheet = workbook[sheet_name]
        else:
            sheet = workbook.active

        headers = {}
        for cell in sheet[1]:
            if cell.value is not None:
                headers[cell.value.strip()] = cell.column_letter

        required_columns = {
            'Attribute Permission:': None,
            'Launch:': None
        }
        for column_name in required_columns:
            if column_name not in headers:
                raise ValueError(f"A required column is missing: {column_name}")
            required_columns[column_name] = headers[column_name]

        attr_permission_col = required_columns['Attribute Permission:']
        launch_col = required_columns['Launch:']

        tests_to_run = []

        max_row = end_row if end_row else sheet.max_row

        for row in range(start_row, max_row + 1):
            attribute_permission = sheet[f"{attr_permission_col}{row}"].value
            launch_value = sheet[f"{launch_col}{row}"].value

            if attribute_permission is None or launch_value is None:
                continue

            if str(launch_value).strip().lower() == 'run':
                tests_to_run.append(str(attribute_permission).strip())

        return tests_to_run
